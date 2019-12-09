import requests
from openpyxl import Workbook
from PyQt5.QtWidgets import QApplication, QWidget, QMessageBox, QPushButton, QLineEdit, QLabel,QComboBox
import sys
# from Code.dict_get import dict1,dict2,dict3
import urllib
import string
import datetime

class Mwindow(QWidget):
    def __init__(self):
        super().__init__()
        self.dict2 = {"螺纹钢": "螺纹钢_:_螺纹钢",
                 "角钢": "角钢_:_角钢",
                 "槽钢": "槽钢_:_槽钢",
                 "工字钢": "工字钢_:_工字钢",
                 "线材": "线材_:_线材",
                 "冷轧板卷": "冷轧板卷_:_冷轧板卷",
                 "造船板": "造船板_:_造船板",
                 "中厚板": "中厚板_:_中厚板",
                 "热轧板卷": "热轧板卷_:_热轧板卷",
                 "无缝管": "无缝管_:_无缝管",
                 "镀锌板卷": "镀锌板卷_:_镀锌板卷",
                 "彩涂板卷": "彩涂板卷_:_彩涂板卷",
                 "热轧带钢": "热轧带钢_:_热轧带钢",
                 "H型钢": "H型钢_:_H型钢",
                 "焊管": "焊管_:_焊管",
                 "硅钢": "硅钢_:_硅钢",
                 "热轧管坯": "热轧管坯_:_热轧管坯",
                 "Cr合结钢": "Cr合结钢_:_Cr合结钢",
                 "碳结圆钢": "碳结圆钢_:_碳结圆钢",
                 "轴承圆钢": "轴承圆钢_:_轴承圆钢",
                 "焊线": "焊线_:_焊线",
                 "拉丝材": "拉丝材_:_拉丝%25",
                 "冷镦钢": "冷镦钢_:_冷镦钢",
                 "弹簧圆钢": "弹簧圆钢_:_弹簧圆钢",
                 "工模具钢": "工模具钢_:_工模具钢",
                 "齿轮钢": "齿轮钢_:_齿轮钢",
                 "Cr系不锈圆钢": "Cr系不锈圆钢_:_Cr系不锈圆钢",
                 "碳结板": "碳结板_:_碳结板",
                 "电解镍": "电解镍_:_电解镍",
                 "电解铜": "电解铜_:_电解铜",
                 "锡锭": "锡锭_:_锡锭",
                 "钨铁": "钨铁_:_钨铁",
                 "电解铝": "电解铝_:_电解铝",
                 "铅锭": "铅锭_:_铅锭",
                 "锌锭": "锌锭_:_锌锭",
                 "ADC12": "ADC12_:_ADC12",
                 "A356.2": "A356.2_:_A356.2",
                 "镁锭": "镁锭_:_镁锭",
                 "锌合金": "锌合金_:_锌合金",
                 "硬线": "硬线_:_硬线",
                 "甲醇": "甲醇_:_甲醇",
                 "工业萘": "工业萘_:_工业萘"}

        self.dict1 = {"螺纹钢": ["HRB335 12MM", "HRB335 20MM", "HRB400 20MM"],
                 "角钢": ["50*5"],
                 "槽钢": ["16#"],
                 "工字钢": ["25#"],
                 "线材": ["6高线HPB300", "8.0高线HPB300"],
                 "冷轧板卷": ["0.5mm_", "1.0mm_"],
                 "造船板": ["10mm", "20mm"],
                 "中厚板": ["普8mm", "普20mm", "低合金20mm"],
                 "热轧板卷": ["3.0", "4.75"],
                 "无缝管": ["108*4.5（8163）", "219*6(8163)", "108*4.5（8162）", "219*6(8162)"],
                 "镀锌板卷": ["0.5mm", "1.0mm"],
                 "彩涂板卷": ["0.476mm"],
                 "热轧带钢": ["3.5mm*685", "2.5mm*232"],
                 "H型钢": ["200*100", "300*300", "400*200", "588*300"],
                 "焊管": ["1.5寸*3.0", "1.5寸*3.25", "4寸*3.75"],
                 "硅钢": ["WW600", "WW800"],
                 "热轧管坯": ["∮50-130mm"],
                 "Cr合结钢": ["Φ20mm_", "Φ80mm_", "Φ140mm_"],
                 "碳结圆钢": ["Φ20mm", "Φ80mm", "Φ140mm"],
                 "轴承圆钢": ["50mm(连铸)"],
                 "焊线": ["H08A"],
                 "拉丝材": ["Φ6.5"],
                 "冷镦钢": ["22A（18A）", "ML08AL（8A）", "ML35（35K）"],
                 "弹簧圆钢": ["60Si2Mn"],
                 "工模具钢": ["Cr12MoV", "3Cr2W8V", "H13(电炉)"],
                 "齿轮钢": ["20CrMnTiΦ50", "20CrMoΦ50", "35/42CrMoΦ50"],
                 "Cr系不锈圆钢": ["2Cr13"],
                 "碳结板": ["50mm"],
                 "锌合金": ["Zamak-5", "Zamak-3", "铸造", "热镀"],
                 "硬线": ["45-70#"],
                 "电解镍": [],
                 "电解铜": [],
                 "锡锭": [],
                 "钨铁": [],
                 "电解铝": [],
                 "铅锭": [],
                 "ADC12": [],
                 "A356.2": [],
                 "镁锭": [],
                 "甲醇": [],
                 "工业萘": []
                 }

        # for values in dict1.values():
        #     for k in values:
        #         print('"'+k+'"'+':')
        #
        # for i in soup.find_all('input',{"name":"specs"}):
        #     print('"'+i['value']+'"')

        self.dict3 = {"HRB335 12MM": "螺纹钢_:_螺纹钢:__:HRB335 12MM_:_HRB335_12MM",
                 "HRB335 20MM": "螺纹钢_:_螺纹钢:__:HRB335 20MM_:_HRB335_20MM",
                 "HRB400 20MM": "螺纹钢_:_螺纹钢:__:HRB400 20MM_:_HRB400_20MM",
                 "50*5": "角钢_:_角钢:__:50*5_:__50*5",
                 "16#": "槽钢_:_槽钢:__:16#_:__16#",
                 "25#": "工字钢_:_工字钢:__:25#_:__25#",
                 "6高线HPB300": "线材_:_线材:__:6高线HPB300_:_HPB300_6高线",
                 "8.0高线HPB300": "线材_:_线材:__:8.0高线HPB300_:_HPB300_8.0高线",
                 "0.5mm_": "冷轧板卷_:_冷轧板卷:__:0.5mm_:__0.5mm",
                 "1.0mm_": "冷轧板卷_:_冷轧板卷:__:1.0mm_:__1.0mm",
                 "10mm": "造船板_:_造船板:__:10mm_:__10mm",
                 "20mm": "造船板_:_造船板:__:20mm_:__20mm",
                 "普8mm": "中厚板_:_中厚板:__:普8mm_:__普8mm",
                 "普20mm": "中厚板_:_中厚板:__:普20mm_:__普20mm",
                 "低合金20mm": "中厚板_:_中厚板:__:低合金20mm_:__低合金20mm",
                 "3.0": "热轧板卷_:_热轧板卷:__:3.0_:__%253.0%25",
                 "4.75": "热轧板卷_:_热轧板卷:__:4.75_:__%254.75%25",
                 "108*4.5（8163）": "无缝管_:_无缝管:__:108*4.5（8163）_:__108*4.5%258163%25",
                 "219*6(8163)": "无缝管_:_无缝管:__:219*6(8163)_:__219*6%258163%25",
                 "108*4.5（8162）": "无缝管_:_无缝管:__:108*4.5（8162）_:__108*4.5%258162%25",
                 "219*6(8162)": "无缝管_:_无缝管:__:219*6(8162)_:__219*6%258162%25",
                 "0.5mm": "镀锌板卷_:_镀锌板卷:__:0.5mm_:__0.5mm",
                 "1.0mm": "镀锌板卷_:_镀锌板卷:__:1.0mm_:__1.0mm",
                 "0.476mm": "彩涂板卷_:_彩涂板卷:__:0.476mm_:__0.476mm",
                 "3.5mm*685": "热轧带钢_:_热轧带钢:__:3.5mm*685_:__3.5mm*685",
                 "2.5mm*232": "热轧带钢_:_热轧带钢:__:2.5mm*232_:__2.5mm*232",
                 "200*100": "H型钢_:_H型钢:__:200*100_:__200*100",
                 "300*300": "H型钢_:_H型钢:__:300*300_:__300*300",
                 "400*200": "H型钢_:_H型钢:__:400*200_:__400*200",
                 "588*300": "H型钢_:_H型钢:__:588*300_:__588*300",
                 "1.5寸*3.0": "焊管_:_焊管:__:1.5寸*3.0_:__1.5寸*3.0",
                 "1.5寸*3.25": "焊管_:_焊管:__:1.5寸*3.25_:__1.5寸*3.25",
                 "4寸*3.75": "焊管_:_焊管:__:4寸*3.75_:__4寸*3.75",
                 "WW600": "硅钢_:_硅钢:__:WW600_:_%25WW600_%25",
                 "WW800": "硅钢_:_硅钢:__:WW800_:_%25WW800_%25",
                 "∮50-130mm": "热轧管坯_:_热轧管坯:__:∮50-130mm_:__∮50-130mm",
                 "Φ20mm_": "Cr合结钢_:_Cr合结钢:__:Φ20mm_:__Φ20mm",
                 "Φ80mm_": "Cr合结钢_:_Cr合结钢:__:Φ80mm_:__Φ80mm",
                 "Φ140mm_": "Cr合结钢_:_Cr合结钢:__:Φ140mm_:__Φ140mm",
                 "Φ20mm": "碳结圆钢_:_碳结圆钢:__:Φ20mm_:__Φ20%25",
                 "Φ80mm": "碳结圆钢_:_碳结圆钢:__:Φ80mm_:__Φ80%25",
                 "Φ140mm": "碳结圆钢_:_碳结圆钢:__:Φ140mm_:__Φ140%25",
                 "50mm(连铸)": "轴承圆钢_:_轴承圆钢:__:50mm(连铸)_:__50mm%25连铸%25",
                 "H08A": "焊线_:_焊线:__:H08A_:_H08A_φ5.5",
                 "Φ6.5": "拉丝材_:_拉丝%25:__:Φ6.5_:__Φ6.5",
                 "22A（18A）": "冷镦钢_:_冷镦钢:__:22A（18A）_:_22A（18A）_%25",
                 "ML08AL（8A）": "冷镦钢_:_冷镦钢:__:ML08AL（8A）_:_ML08AL（8A）_%25",
                 "ML35（35K）": "冷镦钢_:_冷镦钢:__:ML35（35K）_:_ML35（35K）_%25",
                 "60Si2Mn": "弹簧圆钢_:_弹簧圆钢:__:60Si2Mn_:__60Si2Mn",
                 "Cr12MoV": "工模具钢_:_工模具钢:__:Cr12MoV_:_Cr12MoV_%25",
                 "3Cr2W8V": "工模具钢_:_工模具钢:__:3Cr2W8V_:_3Cr2W8V_%25",
                 "H13(电炉)": "工模具钢_:_工模具钢:__:H13(电炉)_:_H13(电炉)_%25",
                 "20CrMnTiΦ50": "齿轮钢_:_齿轮钢:__:20CrMnTiΦ50_:__20CrMnTi%25",
                 "20CrMoΦ50": "齿轮钢_:_齿轮钢:__:20CrMoΦ50_:__20CrMo%25",
                 "35/42CrMoΦ50": "齿轮钢_:_齿轮钢:__:35/42CrMoΦ50_:__35/42CrMo%25",
                 "2Cr13": "Cr系不锈圆钢_:_Cr系不锈圆钢:__:2Cr13_:__2Cr13",
                 "50mm": "碳结板_:_碳结板:__:50mm_:__50mm",
                 "Zamak-5": "锌合金_:_锌合金:__:Zamak-5_:__Zamak-5",
                 "Zamak-3": "锌合金_:_锌合金:__:Zamak-3_:__Zamak-3",
                 "铸造": "锌合金_:_锌合金:__:铸造_:__铸造",
                 "热镀": "锌合金_:_锌合金:__:热镀_:__热镀",
                 "45-70#": "硬线_:_硬线:__:45-70#_:_45-70#_Φ6.5"}



        self.lb1 = QLabel(self)
        self.lb1.setGeometry(100, 200, 41, 31)
        self.lb1.setText('品类：')
        self.lb2 = QLabel(self)
        self.lb2.setGeometry(100, 255, 41, 21)
        self.lb2.setText('型号：')
        self.lb3 = QLabel(self)
        self.lb3.setGeometry(400, 200, 54, 21)
        self.lb3.setText('开始时间：')
        self.lb3 = QLabel(self)
        self.lb3.setGeometry(400, 250, 54, 21)
        self.lb3.setText('结束时间：')


        self.cb1 = QComboBox(self)
        self.cb1.setGeometry(140, 200, 200, 31)
        self.cb1.addItems([x for x in self.dict1.keys()])
        self.cb1.activated.connect(lambda: self.getData(self.cb1.currentText()))

        self.cb2 = QComboBox(self)
        self.cb2.setGeometry(140, 250, 200, 31)

        self.textbox3 = QLineEdit(self)
        self.textbox3.setGeometry(470, 200, 150, 31)
        self.textbox4 = QLineEdit(self)
        self.textbox4.setGeometry(470, 250, 150, 31)

        self.findButton = QPushButton('数据抽取', self)
        self.findButton.setGeometry(330, 330, 75, 33)
        self.findButton.clicked.connect(self.getChartData)


    def getData(self,text):
        self.cb2.clear()
        self.cb2.addItems(self.dict1.get(text,""))

    def getChartData(self):
        url = "https://index.mysteel.com/newprice/getChartMultiCatalog.ms?callback=callback"
        # 请求头、参数设置
        headers = {
            "User-Agent": "PostmanRuntime/7.19.0",
            "Context-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Cookie": "Hm_lvt_1c4432afacfa2301369a5625795031b8=1572335579; href=https%3A%2F%2Findex.mysteel.com%2Fprice%2FindexPrice.html; accessId=5d36a9e0-919c-11e9-903c-ab24dbab411b; JSESSIONID=07332AA24253B7330274CD36B4312418; Hm_lpvt_1c4432afacfa2301369a5625795031b8=1572396557; qimo_seosource_5d36a9e0-919c-11e9-903c-ab24dbab411b=%E7%AB%99%E5%86%85; qimo_seokeywords_5d36a9e0-919c-11e9-903c-ab24dbab411b=; pageViewNum=4"}


        catalog_get = self.cb1.currentText()
        catalog_data = str(self.dict2.get(catalog_get,""))
        print(catalog_data)
        catalog = urllib.parse.quote(catalog_data, safe=string.printable)
        print(catalog)

        spec_get = self.cb2.currentText()
        spec_data = str(self.dict3.get(spec_get,""))
        print(spec_data)
        spec = urllib.parse.quote(spec_data, safe=string.printable)
        print(spec)

        #获取时间
        startTime = str(self.textbox3.text())
        endTime = str(self.textbox4.text())
        print(startTime, endTime)

        data = {"catalog": catalog,
                "city": "%E4%B8%8A%E6%B5%B7",
                "spec": spec,
                "startTime": startTime,
                "endTime": endTime
                }

        try:
            r= requests.post(url=url, data=data, headers=headers,verify=False)
            r.raise_for_status()
            r.encoding = 'utf-8'
            html = r.text
            print(html)
            print('成功')

        except Exception as e:
            QMessageBox.information(self, "提示", "请确认输入是否正确！或网络是否为外网！/n"+str(e),
                                    QMessageBox.Yes)
        try:
        #网页解析
            chart_data = eval(html[9:-2])
            chart_list = chart_data['data']
            date = []
            value = []
            for k in chart_list:   #chart_list为一个列表
                lineNme = k['lineName']   #获得产品线
                # print(lineNme)
                for i in k['dateValueMap']:
                    date.append(i['date'])
                    value.append(i['value'])

        # 写入excel
            wb = Workbook()  # 创建文件对象
            # grab the active worksheet
            ws = wb.active  # 获取第一个sheet
            ws.cell(1, 1).value = "Date"
            ws.cell(1, 2).value = "Price"
            for i in range(2,len(date)+2):
                ws.cell(i,1).value=datetime.datetime.strptime(date[i-2],"%Y-%m-%d").strftime("%Y/%m/%d")

            for k in range(2,len(value)+2):
                ws.cell(k,2).value=('%.2f' % float(value[k-2]))

            # wb.save(r"\\10.56.3.11\566服务中心\08_IT工具"+catalog_data+"-"+spec_data.replace("*"," ")+datetime.datetime.strptime(endTime,"%Y-%m-%d").strftime("%y%m%d")+".xlsx")
            wb.save("e:\\" + catalog_get + "-" + spec_get.replace("*", " ") + datetime.datetime.strptime(endTime, "%Y-%m-%d").strftime("%y%m%d") + ".xlsx")
            QMessageBox.information(self, "提示", "数据抽取成功！",
                                    QMessageBox.Yes)

        except Exception as e:
            QMessageBox.information(self, "提示", "数据抽取失败！,请确认是否关闭EXCEL文件!/n"+str(e),
                                    QMessageBox.Yes)



if __name__ == '__main__':
    app = QApplication(sys.argv)
    find_weather = Mwindow()
    find_weather.show()
    sys.exit(app.exec())











